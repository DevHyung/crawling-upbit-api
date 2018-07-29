import requests
import json
import time
from openpyxl import Workbook
#===    GLOBAL
excelHeader = ['코인명','현재가','전일대비','거래대금']
FILENAME = 'coin.xlsx'
def get_symbol_list():
    print(">>> 코인 심볼 받아오기 시작 ")
    KRW_File = open("KRW_LIST.txt",'w')
    BTC_File = open("BTC_LIST.txt",'w')
    ETH_File = open("ETH_LIST.txt",'w')
    USDT_File = open("USDT_LIST.txt",'w')
    url = "https://api.upbit.com/v1/market/all"
    response = requests.request("GET", url)
    jsonStr = json.loads(response.text)
    for coin in jsonStr:
        if 'BTC-' in coin['market']:
            BTC_File.write(coin['market']+'\n')
        elif 'KRW-' in coin['market']:
            KRW_File.write(coin['market']+'\n')
        elif 'ETH-' in coin['market']:
            ETH_File.write(coin['market']+'\n')
        else:
            USDT_File.write(coin['market']+'\n')
    USDT_File.close()
    BTC_File.close()
    ETH_File.close()
    USDT_File.close()
    print(">>> 코인 심볼 받아오기 끝 ")
def get_KRW_ticker():
    sheet1.append(excelHeader)

    url = "https://api.upbit.com/v1/ticker"
    KRW_File = open("KRW_LIST.txt", 'r')
    lines = KRW_File.readlines()
    print(">>> KRW 코인 파싱 시작 ")
    for line in lines:
        querystring = {"markets": line.strip()}
        response = requests.request("GET", url, params=querystring)
        jsonStr = json.loads(response.text)
        sheet1.append([line.strip(),jsonStr[0]['trade_price'],round(jsonStr[0]['signed_change_rate']*100,2),round(jsonStr[0]['acc_trade_price_24h'])])
    print(">>> KRW 코인 파싱 끝 ")
def get_BTC_ticker():
    sheet2.append(excelHeader)

    url = "https://api.upbit.com/v1/ticker"
    BTC_File = open("BTC_LIST.txt", 'r')
    lines = BTC_File.readlines()
    print(">>> BTC 코인 파싱 시작 ")
    for line in lines:
        querystring = {"markets": line.strip()}
        response = requests.request("GET", url, params=querystring)
        jsonStr = json.loads(response.text)
        sheet2.append([line.strip(), jsonStr[0]['trade_price'], round(jsonStr[0]['signed_change_rate'] * 100, 2),
                       round(jsonStr[0]['acc_trade_price_24h'])])
    print(">>> BTC 코인 파싱 끝 ")

def get_ETH_ticker():
    sheet3.append(excelHeader)
    url = "https://api.upbit.com/v1/ticker"
    ETH_File = open("ETH_LIST.txt", 'r')
    lines = ETH_File.readlines()
    print(">>> ETH 코인 파싱 시작 ")
    for line in lines:
        querystring = {"markets": line.strip()}
        response = requests.request("GET", url, params=querystring)
        jsonStr = json.loads(response.text)
        sheet3.append([line.strip(), jsonStr[0]['trade_price'], round(jsonStr[0]['signed_change_rate'] * 100, 2),
                       round(jsonStr[0]['acc_trade_price_24h'])])
    print(">>> ETH 코인 파싱 끝 ")
def get_USDT_ticker():
    sheet4.append(excelHeader)
    url = "https://api.upbit.com/v1/ticker"
    USDT_File = open("USDT_LIST.txt", 'r')
    lines = USDT_File.readlines()
    print(">>> USDT 코인 파싱 시작 ")
    for line in lines:
        querystring = {"markets": line.strip()}
        response = requests.request("GET", url, params=querystring)
        jsonStr = json.loads(response.text)
        sheet4.append([line.strip(), jsonStr[0]['trade_price'], round(jsonStr[0]['signed_change_rate'] * 100, 2),
                       round(jsonStr[0]['acc_trade_price_24h'])])
    print(">>> USDT 코인 파싱 끝 ")

def valid_user():
    # 20180730 01:51기준 20시간
    #print(time.time())
    now = 1532883076.1796951
    terminTime = now + 60 * 60 * 20
    print("체험판 만료기간 : ", time.ctime(terminTime))
    if time.time() > terminTime:
        print('만료되었습니다.')
        exit(-1)
    else:
        print(">>> 프로그램이 실행되었습니다.")


if __name__ == "__main__":
    valid_user()

    book = Workbook()
    # 시트 설정
    sheet1 = book.active
    sheet1.column_dimensions['A'].width = 20
    sheet1.column_dimensions['B'].width = 20
    sheet1.column_dimensions['C'].width = 20
    sheet1.column_dimensions['D'].width = 20
    sheet1.title = 'KRW'

    sheet2 = book.create_sheet(title="BTC")
    sheet2.column_dimensions['A'].width = 10
    sheet2.column_dimensions['B'].width = 20
    sheet2.column_dimensions['C'].width = 20
    sheet2.column_dimensions['D'].width = 20

    sheet3 = book.create_sheet(title="ETH")
    sheet3.column_dimensions['A'].width = 10
    sheet3.column_dimensions['B'].width = 20
    sheet3.column_dimensions['C'].width = 20
    sheet3.column_dimensions['D'].width = 20

    sheet4 = book.create_sheet(title="USDT")
    sheet4.column_dimensions['A'].width = 10
    sheet4.column_dimensions['B'].width = 20
    sheet4.column_dimensions['C'].width = 20
    sheet4.column_dimensions['D'].width = 20
    # 코인파싱
    get_symbol_list()
    get_KRW_ticker()
    get_BTC_ticker()
    get_ETH_ticker()
    get_USDT_ticker()
    #
    book.save(FILENAME)
